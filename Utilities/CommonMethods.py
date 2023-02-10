import datetime
import os

import allure
import openpyxl
import requests
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait

from Utilities import readProperties as config


class CommonMethods:
    @staticmethod
    def getWebElement(self, locators_section, locator_xpath):
        wait = WebDriverWait(self.driver, 30)
        elem_ = wait.until(
            ec.presence_of_element_located((By.XPATH, config.readLocator(locators_section, locator_xpath))))
        return elem_

    @staticmethod
    def getWebElements(self, locators_section, locator_xpath):
        wait = WebDriverWait(self.driver, 30)
        elem_ = wait.until(
            ec.presence_of_all_elements_located((By.XPATH, config.readLocator(locators_section, locator_xpath))))
        return elem_

    @staticmethod
    def validatePageTitle(self, _expectedPageTitle, _testCaseName):
        page_title = self.driver.title
        if page_title == _expectedPageTitle:
            assert True
        else:
            print(page_title)
            CommonMethods.captureScreenshot(self, _testCaseName)
            assert False

    @staticmethod
    def captureScreenshot(self, testCaseName):
        allure.attach(self.driver.get_screenshot_as_png(), name=testCaseName,
                      attachment_type=AttachmentType.PNG)

    @staticmethod
    def findBrokenLinks(self):
        all_links = self.driver.find_elements(By.XPATH, "//a")
        for link in all_links:
            url = link.get_attribute('href')
            result = requests.head(url)
            """
            if status code is not 200 then print the url 
            """
            if result.status_code != 200:
                print(url, result.status_code)

    @staticmethod
    def scrollToElement(self, element_locator):
        self.driver.execute_script("arguments[0].scrollIntoView();", element_locator)

    @staticmethod
    def scroll_page(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    @staticmethod
    def clickUsingScript(self, element):
        self.driver.execute_script('arguments[0].click()', element)

    @staticmethod
    def fetchTestDataRowWise(worksheet_name, rowStartIndexDataStartReading, colFromWhereDataStartReading, dataFilePath):
        workbook = openpyxl.load_workbook(dataFilePath)
        worksheet = workbook[worksheet_name]
        rows = worksheet.max_row
        cols = worksheet.max_column
        # print(f"Total number of rows {rows} and Total Number of columns are {cols}")
        dataList = []
        for i in range(rowStartIndexDataStartReading, rows + 1):
            tempDataList = []
            for j in range(colFromWhereDataStartReading, cols + 1):
                data1 = worksheet.cell(i, j)
                data2 = data1.value
                if data2 is not None:
                    tempDataList.append(data2)
            dataList.insert(i, tempDataList)
        return dataList

    # print(CommonMethods.fetchTestDataRowWise('SignUpPage', 2, 1, r"../TestData/TestData.xlsx"))

    @staticmethod
    def fileUpload(basePath, fileNameWithType):
        path_to_image = os.path.join(basePath, fileNameWithType)
        return path_to_image

    @staticmethod
    def dateFormatter(user_date):
        specified_datetime = datetime.datetime.strptime(user_date, "%Y-%m-%d")
        formatted_date = specified_datetime.strftime("%m/%d/%Y")
        return formatted_date
